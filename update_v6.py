"""
update_v6.py
============
在庫一覧.csv と 最新の *実績.csv を読み込み、
Kanken23510_2026年版_v6.xlsx の「全体サマリー」シートを自動更新する。

【更新対象列】
  D列: 倉庫WH    ← バルク + New Way-A の合計現在数量
  E列: 店舗在庫  ← 全実店舗 + EC（ZOZO/丸井）の合計現在数量
  I列: 需要予測  ← 直近実績ベースで再計算（スケール係数を乗じる）
  L列: 過不足    ← H列(総供給) - I列(需要予測) で再計算
"""

import sys
import os
import glob
import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')

# ==============================================================
# 設定エリア
# ==============================================================
EXCEL_FILE  = 'Kanken23510_2026年版_v6.xlsx'
STOCK_CSV   = '在庫一覧.csv'
BUNPAI_XLSX = '分配店舗向け.xlsx'

# 倉庫WH として集計する拠点名（現在数量の合計をD列に書く）
WAREHOUSE_LOCATIONS = {
    'バルク',
    'New Way-A',
}

# 除外する拠点名（不良在庫・集計対象外）
EXCLUDE_LOCATIONS = {
    'New Way-B',
    'New Way-C',
    'New Way-G',
}
# ※上記以外の全拠点 = 店舗在庫（ZOZO・丸井ｳｪﾌﾞも含む）

# 2024年実績ベースの月別累積進捗率（backend_calc.py と共通）
CUMULATIVE_PROGRESS = {
    1: 0.0437, 2: 0.1050, 3: 0.1970, 4: 0.3090, 5: 0.3968, 6: 0.4921,
    7: 0.5968, 8: 0.7067, 9: 0.7903, 10: 0.8677, 11: 0.9360, 12: 1.0000
}

MONTH_MAP = {
    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5,  'JUN': 6,
    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
}

# v6「全体サマリー」の列番号（1始まり・openpyxl基準）
COL_COLOR  = 3   # C: カラー
COL_WH     = 4   # D: 倉庫WH
COL_STORE  = 5   # E: 店舗在庫
COL_SUPPLY = 8   # H: 総供給(6月〜)
COL_DEMAND = 9   # I: 需要予測(6〜12月)
COL_DIFF   = 12  # L: 過不足

HEADER_ROW = 4  # 列ヘッダー行
DATA_START = 5  # データ開始行
# ==============================================================


def find_latest_sales_csv() -> str:
    """フォルダ内の *実績.csv を探し、最新の1ファイルを返す。"""
    base_dir   = os.path.dirname(os.path.abspath(__file__))
    pattern    = os.path.join(base_dir, '*実績.csv')
    candidates = glob.glob(pattern)
    if not candidates:
        raise FileNotFoundError(
            "'*実績.csv' が見つかりません。フォルダに実績CSVを配置してください。"
        )
    latest = max(candidates, key=os.path.getmtime)
    return latest


def load_master(path: str) -> dict:
    """分配店舗向け.xlsx の master シートから {品番: カラー名} を返す。"""
    df = pd.read_excel(path, sheet_name='master')
    df['品番']    = df['品番'].astype(str).str.strip()
    df['カラー名'] = df['カラー名'].astype(str).str.strip()
    return dict(zip(df['品番'], df['カラー名']))


def load_stock(path: str, master: dict) -> tuple:
    """
    在庫一覧.csv から カラー別の (倉庫WH合計, 店舗在庫合計) を返す。
    Returns:
        wh_stock   : {カラー名: 倉庫WH合計}
        store_stock: {カラー名: 店舗在庫合計}
    """
    try:
        df = pd.read_csv(path, encoding='cp932')
    except UnicodeDecodeError:
        df = pd.read_csv(path, encoding='utf-8')

    df['商品コード'] = df['商品コード'].astype(str).str.strip()
    df['現在数量']   = pd.to_numeric(df['現在数量'], errors='coerce').fillna(0).clip(lower=0)

    # カラー名を付与
    df['カラー名'] = df['商品コード'].map(master)

    # 紐付けできなかった行を警告表示
    unmapped = df[df['カラー名'].isna()]['商品コード'].unique()
    if len(unmapped) > 0:
        print(f"  [警告] masterに存在しない商品コード（スキップ）: {unmapped.tolist()}")

    df = df.dropna(subset=['カラー名'])

    # 除外拠点を除く
    df = df[~df['拠点名'].isin(EXCLUDE_LOCATIONS)]

    # 倉庫WH と 店舗在庫 に分類
    df['区分'] = df['拠点名'].apply(
        lambda x: '倉庫' if x in WAREHOUSE_LOCATIONS else '店舗'
    )

    wh_stock    = df[df['区分'] == '倉庫'].groupby('カラー名')['現在数量'].sum().astype(int).to_dict()
    store_stock = df[df['区分'] == '店舗'].groupby('カラー名')['現在数量'].sum().astype(int).to_dict()

    return wh_stock, store_stock


def load_sales(path: str, master: dict) -> tuple:
    """
    実績CSVからカラー別販売数の合計と実績月数を返す。
    Returns:
        sales_by_color: {カラー名: 販売数合計}
        months_count  : 実績に含まれる月数
    """
    try:
        df = pd.read_csv(path, encoding='cp932')
    except UnicodeDecodeError:
        df = pd.read_csv(path, encoding='utf-8')

    # 列名のゆらぎ対応
    if '販売数' in df.columns:
        qty_col = '販売数'
    elif '数量' in df.columns:
        qty_col = '数量'
    else:
        raise ValueError("実績CSVに '販売数' または '数量' 列が見つかりません。")

    df[qty_col] = pd.to_numeric(df[qty_col], errors='coerce').fillna(0).clip(lower=0)

    # 品番または商品コード列の探索
    code_col = None
    for col in ['商品コード', '品番', '行ラベル']:
        if col in df.columns:
            code_col = col
            break

    if code_col is None:
        raise ValueError("実績CSVに商品コードまたは品番を表す列が見つかりません。")

    df['コードキー'] = df[code_col].astype(str).str.strip()
    
    # マスタでカラー名に変換
    df['カラー名_マッピング済'] = df['コードキー'].map(master)

    # Kanken(23510)以外のレコードも含まれるため、マッピングできた行のみを集計対象とする
    kanken_df = df.dropna(subset=['カラー名_マッピング済'])

    sales_by_color = kanken_df.groupby('カラー名_マッピング済')[qty_col].sum().astype(int).to_dict()

    # 実績月数の算定（数量合計が1以上の月をカウント）
    if 'MTH' in df.columns:
        monthly_sales = df.groupby('MTH')[qty_col].sum()
        valid_months = [m for m, val in monthly_sales.items() if val > 0 and str(m).upper() in MONTH_MAP]
        months_count = len(valid_months)
    else:
        months_count = 1

    months_count = max(1, min(12, months_count))
    return sales_by_color, months_count


def calc_scale_factor(sales_by_color: dict, months_count: int, ws) -> tuple:
    """
    実績ベースの需要予測スケール係数を算出する。
    「全カラー合計の実績 / 進捗率 = 年間予測」→「残り月比率」で6〜12月需要を推計。
    Returns: (scale, annual_forecast, new_6to12_demand)
    """
    progress_rate   = CUMULATIVE_PROGRESS.get(months_count, 1.0)
    total_actual    = sum(sales_by_color.values())
    annual_forecast = total_actual / progress_rate
    future_rate     = 1.0 - progress_rate
    new_6to12_demand = annual_forecast * future_rate

    # 現行v6の需要予測合計を取得（スケール基準）
    old_demand_total = 0.0
    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row):
        color_cell  = row[COL_COLOR  - 1]
        demand_cell = row[COL_DEMAND - 1]
        if color_cell.value is None or str(color_cell.value).strip() in ('', '合計'):
            continue
        if isinstance(demand_cell.value, (int, float)):
            old_demand_total += float(demand_cell.value)

    if old_demand_total == 0:
        print("  [情報] v6エクセルの旧需要予測合計がゼロです。実績ベースの直接算出に切り替えます。")
        scale = None
    else:
        scale = new_6to12_demand / old_demand_total
    return scale, annual_forecast, new_6to12_demand


def main():
    print("=" * 60)
    print("  v6 データ自動更新スクリプト")
    print("=" * 60)

    # --- 1. マスタ読込 ---
    print("\n[1/5] マスタデータ（品番↔カラー名）を読み込み中...")
    master = load_master(BUNPAI_XLSX)
    print(f"      {len(master)} 品番を読み込みました。")

    # --- 2. 在庫一覧読込 ---
    print(f"\n[2/5] 在庫一覧 ({STOCK_CSV}) を読み込み中...")
    wh_stock, store_stock = load_stock(STOCK_CSV, master)
    print(f"      倉庫WH: {len(wh_stock)} カラー / 店舗在庫: {len(store_stock)} カラー")
    print(f"      倉庫WH 合計: {sum(wh_stock.values()):,}個 / 店舗在庫 合計: {sum(store_stock.values()):,}個")

    # --- 3. 実績CSV読込 ---
    print(f"\n[3/5] 実績CSVを自動検索・読み込み中...")
    sales_path = find_latest_sales_csv()
    print(f"      使用ファイル: {os.path.basename(sales_path)}")
    sales_by_color, months_count = load_sales(sales_path, master)
    progress_rate = CUMULATIVE_PROGRESS.get(months_count, 1.0)
    print(f"      実績月数: {months_count} ヶ月 (進捗率: {progress_rate*100:.2f}%)")
    print(f"      合計実績: {sum(sales_by_color.values()):,}個")

    # --- 4. v6読込・スケール係数計算 ---
    print(f"\n[4/5] v6ファイルを読み込み・スケール係数を計算中...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['全体サマリー']

    scale, annual_fc, new_demand_total = calc_scale_factor(sales_by_color, months_count, ws)
    print(f"      年間需要予測(実績ベース): {annual_fc:,.0f}個")
    print(f"      6〜12月 新需要予測合計 : {new_demand_total:,.0f}個")
    if scale is not None:
        print(f"      スケール係数           : {scale:.4f}  ({(scale-1)*100:+.1f}%)")
    else:
        print(f"      スケール係数           : 適用不可 (エクセル側需要予測が0のためダイレクト計算を実行)")

    # --- 5. v6へ書き込み ---
    print(f"\n[5/5] v6「全体サマリー」シートを更新中...")
    print(f"\n  {'カラー':<30} {'旧倉庫WH':>8} {'新倉庫WH':>8} {'旧店舗':>7} {'新店舗':>7} {'旧需要':>7} {'新需要':>7} {'新過不足':>9}")
    print("  " + "-" * 93)

    updated = 0
    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row):
        color_cell  = row[COL_COLOR  - 1]
        wh_cell     = row[COL_WH     - 1]
        store_cell  = row[COL_STORE  - 1]
        supply_cell = row[COL_SUPPLY - 1]
        demand_cell = row[COL_DEMAND - 1]
        diff_cell   = row[COL_DIFF   - 1]

        color = str(color_cell.value).strip() if color_cell.value else ''
        if not color or color == '合計':
            continue
        if not isinstance(demand_cell.value, (int, float)):
            continue

        old_wh     = int(wh_cell.value)    if isinstance(wh_cell.value,    (int, float)) else 0
        old_store  = int(store_cell.value)  if isinstance(store_cell.value,  (int, float)) else 0
        old_demand = float(demand_cell.value)
        supply     = float(supply_cell.value) if isinstance(supply_cell.value, (int, float)) else 0.0

        # 在庫データにない場合は旧値を保持し、警告を出す
        if color in wh_stock:
            new_wh = wh_stock[color]
        else:
            new_wh = old_wh
            print(f"  [情報] '{color}': 在庫一覧に倉庫データなし → 旧値 {old_wh:,} を保持")

        if color in store_stock:
            new_store = store_stock[color]
        else:
            new_store = old_store
            print(f"  [情報] '{color}': 在庫一覧に店舗データなし → 旧値 {old_store:,} を保持")

        if scale is not None:
            new_demand = round(old_demand * scale)
        else:
            # エクセルの需要予測がゼロだった場合のフォールバック（実績からダイレクト計算）
            color_sales = sales_by_color.get(color, 0)
            if progress_rate < 1.0 and progress_rate > 0:
                new_demand = round(color_sales * ((1.0 - progress_rate) / progress_rate))
            else:
                new_demand = 0

        new_diff   = round(supply - new_demand)

        print(f"  {color:<30} {old_wh:>8,} {new_wh:>8,} {old_store:>7,} {new_store:>7,} {old_demand:>7,.0f} {new_demand:>7,} {new_diff:>+9,}")

        wh_cell.value     = new_wh
        store_cell.value  = new_store
        demand_cell.value = new_demand
        diff_cell.value   = new_diff
        updated += 1

    wb.save(EXCEL_FILE)
    print("  " + "-" * 93)
    print(f"\n✅ {updated} カラーを更新し、{EXCEL_FILE} を上書き保存しました。")
    print("=" * 60)


if __name__ == '__main__':
    main()
