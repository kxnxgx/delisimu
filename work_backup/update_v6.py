"""
Kanken23510 v5 -> v6 更新スクリプト
- 6月実績を 2026実績.csv に追記
- 指定カラーのトレンド比を163%ベースに更新し v6.xlsx を生成
- backend_calc.py を v6 参照に変更して再実行
"""
import openpyxl
import shutil
import csv
import re
import os
import subprocess
import sys
import io
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = r"c:\Users\kxnxg\OneDrive\デスクトップ\シミュ2"
FILE_V5      = os.path.join(BASE, "Kanken23510_2026年版_v5.xlsx")
FILE_V6      = os.path.join(BASE, "Kanken23510_2026年版_v6.xlsx")
FILE_JISSEKI = os.path.join(BASE, "6.24までの実績.xlsx")
FILE_CSV     = os.path.join(BASE, "2026実績.csv")
FILE_BACKEND = os.path.join(BASE, "backend_calc.py")

NEW_TREND = 1.63

COLORS_TO_UPDATE = {
    "Corn":                          1.775,
    "Pastel Lavender-Confetti Pat":  4.000,
    "Chalk Rose":                    2.222,
    "Green":                         2.222,
    "Black Oak":                     2.222,
    "Teal":                          2.222,
    "Fossil":                        2.222,
    "Midnight Purple":               2.222,
}

STORE_MAP = {
    "FJALLRAVEN by 3NITY TOKYO":          "TOKYO",
    "FJALLRAVEN by 3NITY ルクア大阪":      "ルクア大阪",
    "FJALLRAVEN POPUP NARITA":            "FLAGS",
    "FJALLRAVEN by 3NITY 京王新宿":        "京王新宿",
    "FJALLRAVEN by 3NITY 大丸心斎橋":     "大丸心斎橋",
    "FJALLRAVEN by 3NITY玉川高島屋S・C":  "玉川高島屋",
    "FJALLRAVEN STORE 名古屋ファッションワン": "名古屋",
    "FJALLRAVEN by 3NITY SAPPORO HUTTE":  "HUTTE",
    "FJALLRAVEN by 3NITY POPUP":          "OIOI",
}

def n(val, default=0):
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", ""))
    except:
        return float(default)

def parse_b2(text):
    r = {"wh0": 0, "eb": 0, "ab": 0, "supply": 0}
    if not text: return r
    m = re.search(r"6/1倉庫\(BULK\+NWA\):\s*(\d+)", text)
    if m: r["wh0"] = int(m.group(1))
    m = re.search(r"EB:(\d+)", text)
    if m: r["eb"] = int(m.group(1))
    m = re.search(r"AB:(\d+)", text)
    if m: r["ab"] = int(m.group(1))
    m = re.search(r"総供給:\s*(\d+)", text)
    if m: r["supply"] = int(m.group(1))
    return r

def global_prop_plan(demand_dict, supply):
    months  = sorted(demand_dict)
    total_d = sum(demand_dict.values())
    if total_d <= 0:  return {m: 0 for m in months}
    if supply >= total_d: return {m: round(v) for m, v in demand_dict.items()}
    plan, alloc = {}, 0
    for i, m in enumerate(months):
        if i == len(months) - 1:
            plan[m] = max(0, supply - alloc)
        else:
            p = round(demand_dict[m] / total_d * supply)
            plan[m] = p; alloc += p
    return plan

def wh_movement(wh0, eb, ab, plan_dict):
    fw   = {7: eb, 9: ab}
    rows, cur = [], wh0
    for m in range(6, 13):
        arr = fw.get(m, 0)
        p   = plan_dict.get(m, 0)
        end = max(0, cur + arr - p)
        rows.append((m, cur, arr, p, end)); cur = end
    return rows

# ---- Phase 1 ----
def process_june_actuals():
    print("=" * 50)
    print("Phase 1: 6月実績を 2026実績.csv に追記")
    print("=" * 50)
    wb = openpyxl.load_workbook(FILE_JISSEKI)
    ws = wb["Sheet2"]
    new_rows, skipped = [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        store_raw = str(row[0]).strip() if row[0] is not None else ""
        sku       = str(row[1]).strip() if row[1] is not None else ""
        hinmei    = str(row[3]).strip() if row[3] is not None else "Kanken"
        color_raw = row[4]
        qty       = n(row[6])
        if qty <= 0: continue
        if store_raw in STORE_MAP:
            tenpo = STORE_MAP[store_raw]
        else:
            try:   tenpo = str(int(float(store_raw)))
            except: skipped.append(store_raw); continue
        color_code = str(int(color_raw)) if isinstance(color_raw, float) else (str(color_raw).strip() if color_raw else "")
        new_rows.append({"YEAR":"2026","MTH":"JUN","拠点":tenpo,"品番コード":sku,"数量":str(int(qty)),"販売合計":"","BRAND":"FRV","品番":"23510","品名":hinmei,"カラー":color_code,"サイズ":"X"})
    if skipped: print(f"  ⚠️ スキップ店舗: {list(set(skipped))}")
    print(f"  追加: {len(new_rows)} 件 / {int(sum(n(r['数量']) for r in new_rows))} 個")
    existing, fieldnames = [], None
    for enc in ["utf-8","utf-8-sig","cp932"]:
        try:
            with open(FILE_CSV, encoding=enc, newline="") as f:
                reader = csv.DictReader(f); fieldnames = reader.fieldnames; existing = list(reader)
            break
        except: continue
    if fieldnames is None: print("  ❌ 既存 CSV 読み込み失敗"); return
    all_rows = existing + new_rows
    with open(FILE_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames); w.writeheader(); w.writerows(all_rows)
    print(f"  ✅ 完了: {len(existing)} + {len(new_rows)} = {len(all_rows)} 件")

# ---- Phase 2 ----
def update_color_sheet(ws, color_name, old_trend):
    scale = NEW_TREND / old_trend
    b2_orig = ws["B2"].value or ""
    b2_new  = re.sub(r"(トレンド比:\s*)[\d.]+", rf"\g<1>{NEW_TREND:.3f}", b2_orig)
    b2_new  = re.sub(r"（[^）]*）", "（全体163%ベース統一）", b2_new, count=1)
    ws["B2"].value = b2_new
    info   = parse_b2(b2_orig)
    supply = info["supply"]
    old_d, old_p = {}, {}
    for ri in range(6, 18):
        m = ri - 5
        old_d[m] = n(ws.cell(ri, 6).value)
        old_p[m] = n(ws.cell(ri, 8).value)
    nd       = {m: old_d[m] * scale for m in range(6, 13)}
    total_nd = sum(nd.values())
    is_con   = supply < total_nd
    np_      = global_prop_plan(nd, supply) if is_con else {m: round(nd[m]) for m in range(6, 13)}
    method   = "区間比例配分" if is_con else "需要追随（制約なし）"
    plan15   = sum(old_p[m] for m in range(1, 6))
    total_np = sum(np_.values())
    annual   = plan15 + total_np
    demand15 = sum(old_d[m] for m in range(1, 6))
    for m in range(6, 13):
        ri = m + 5
        ws.cell(ri, 6).value = round(nd[m])
        ws.cell(ri, 8).value = np_[m]
        ws.cell(ri, 9).value = round(np_[m] / annual * 100, 2) if annual > 0 else 0
    ws.cell(18, 6).value = round(demand15 + total_nd)
    ws.cell(18, 8).value = round(annual)
    ws.cell(18, 9).value = 100
    for i, (month, start, arr, plan, end) in enumerate(wh_movement(info["wh0"], info["eb"], info["ab"], np_)):
        ri = 23 + i
        ws.cell(ri, 3).value = start; ws.cell(ri, 4).value = arr
        ws.cell(ri, 5).value = plan;  ws.cell(ri, 6).value = end
    b19 = ws["B19"].value
    if b19 is not None:
        deficit = supply - round(total_nd)
        ws["B19"].value = (f"📦 供給 {supply:,}個　｜　需要(6〜12月) {round(total_nd):,}個　｜　"
                           + (f"不足(区間比例配分でカバー): {abs(deficit):,}個" if is_con else f"超過(需要追随): {deficit:,}個"))
    return {"supply":supply,"method":method,
            "old_d6_12":{m:old_d[m] for m in range(6,13)},
            "new_d6_12":{m:round(nd[m]) for m in range(6,13)},
            "old_p6_12":{m:old_p[m] for m in range(6,13)},
            "new_p6_12":np_,
            "annual":round(annual),"total_nd":round(total_nd),"is_con":is_con}

def update_zentai(wb, results):
    ws = wb["全体サマリー"]
    for cn, res in results.items():
        tgt = None
        for ri in range(5, 35):
            if ws.cell(ri,3).value and str(ws.cell(ri,3).value).strip() == cn:
                tgt = ri; break
        if tgt is None: print(f"  ⚠️ 全体サマリーに '{cn}' 未発見"); continue
        ws.cell(tgt, 9).value  = res["total_nd"]
        ws.cell(tgt, 10).value = res["method"]
        ws.cell(tgt, 11).value = res["annual"]
        ws.cell(tgt, 12).value = res["supply"] - res["total_nd"]
    for col in [9, 11, 12]:
        ws.cell(32, col).value = round(sum(n(ws.cell(r, col).value) for r in range(5, 32)))

def update_forecast(wb, dd, dp):
    ws = wb["2026年着地予測"]
    cumul = n(ws.cell(9, 8).value)  # MAY 累計
    for m in range(6, 13):
        ri = m + 4
        ws.cell(ri, 4).value = round(n(ws.cell(ri, 4).value) + dd[m])
        new_p = round(n(ws.cell(ri, 6).value) + dp[m])
        ws.cell(ri, 6).value = new_p; cumul += new_p
        ws.cell(ri, 8).value = round(cumul)
    total_d = round(sum(n(ws.cell(r, 4).value) for r in range(5, 17)))
    total_p = round(sum(n(ws.cell(r, 6).value) for r in range(5, 17)))
    ws.cell(17, 4).value = total_d; ws.cell(17, 6).value = total_p
    ws.cell(17, 7).value = 100;    ws.cell(17, 8).value = total_p
    for ri in range(5, 17):
        p = n(ws.cell(ri, 6).value)
        ws.cell(ri, 7).value = round(p / total_p * 100, 2) if total_p else 0

def update_sim(wb, dp):
    ws  = wb["供給需要シミュレーション"]
    cur = n(ws.cell(5, 3).value)
    for m in range(6, 13):
        ri = m - 1
        ws.cell(ri, 3).value = round(cur)
        arr = n(ws.cell(ri, 4).value)
        new_p = max(0, round(n(ws.cell(ri, 5).value) + dp[m]))
        ws.cell(ri, 5).value = new_p
        end = max(0, cur + arr - new_p)
        ws.cell(ri, 6).value = round(end); cur = end
    ws.cell(12, 5).value = round(sum(n(ws.cell(r, 5).value) for r in range(5, 12)))
    ws.cell(12, 6).value = round(cur)

def update_logic(wb):
    ws = wb["計算ロジック説明"]
    for ri in range(1, 60):
        for ci in range(1, 12):
            v = ws.cell(ri, ci).value
            if isinstance(v, str) and ("2.2222" in v or "2.222" in v):
                ws.cell(ri, ci).value = v.replace("2.2222","1.630").replace("2.222","1.630")

def create_v6():
    print("\n" + "=" * 50)
    print("Phase 2: v6 Excel 作成")
    print("=" * 50)
    shutil.copy(FILE_V5, FILE_V6); print(f"  コピー: {os.path.basename(FILE_V5)} → {os.path.basename(FILE_V6)}")
    wb = openpyxl.load_workbook(FILE_V6)
    results = {}
    for color, old_trend in COLORS_TO_UPDATE.items():
        if color not in wb.sheetnames: print(f"  ⚠️ シート未発見: {color}"); continue
        res = update_color_sheet(wb[color], color, old_trend)
        results[color] = res
        dc = res["total_nd"] - round(sum(res["old_d6_12"].values()))
        pc = sum(res["new_p6_12"].values()) - round(sum(res["old_p6_12"].values()))
        print(f"  ✅ {color}: 需要 {dc:+d} / 計画 {pc:+d} ({'制約あり' if res['is_con'] else '需要追随'})")
    dd = {m: sum(res["new_d6_12"][m] - res["old_d6_12"][m] for res in results.values()) for m in range(6,13)}
    dp = {m: sum(res["new_p6_12"][m] - res["old_p6_12"][m] for res in results.values()) for m in range(6,13)}
    print("\n  月別 計画デルタ:")
    for m in range(6,13): print(f"    {m}月: {dp[m]:+d}")
    update_zentai(wb, results)
    update_forecast(wb, dd, dp)
    update_sim(wb, dp)
    update_logic(wb)
    wb.save(FILE_V6)
    print(f"\n  ✅ v6 保存: {os.path.basename(FILE_V6)}")

# ---- Phase 3 ----
def update_and_run_backend():
    print("\n" + "=" * 50)
    print("Phase 3: ダッシュボード更新")
    print("=" * 50)
    with open(FILE_BACKEND, "r", encoding="utf-8") as f: content = f.read()
    if "v5.xlsx" in content:
        content = content.replace("v5.xlsx", "v6.xlsx")
        with open(FILE_BACKEND, "w", encoding="utf-8") as f: f.write(content)
        print("  ✅ backend_calc.py: v5 → v6")
    else: print("  ℹ️ すでに v6 参照")
    result = subprocess.run([sys.executable, FILE_BACKEND], cwd=BASE, capture_output=True, text=True, encoding="utf-8", errors="replace")
    for line in result.stdout.strip().split("\n"): print(f"  > {line}")
    if result.returncode != 0: print(f"  ⚠️ エラー: {result.stderr[:500]}")
    else: print("  ✅ ダッシュボード更新完了")

def main():
    print("\n" + "=" * 60)
    print("  Kanken23510 v5 → v6 更新スクリプト 開始")
    print("=" * 60)
    errors = []
    for label, fn in [("Phase1", process_june_actuals), ("Phase2", create_v6), ("Phase3", update_and_run_backend)]:
        try: fn()
        except Exception as e:
            import traceback; errors.append(f"{label}: {e}"); traceback.print_exc()
    print("\n" + "=" * 60)
    if errors: print(f"  ⚠️ 完了（エラー {len(errors)} 件）"); [print(f"    - {e}") for e in errors]
    else:       print("  ✅ 全処理完了（エラーなし）")
    print("=" * 60)

if __name__ == "__main__": main()
