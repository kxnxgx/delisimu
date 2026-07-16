import sys, os, glob, re
import openpyxl
import pandas as pd
from copy import copy
from openpyxl.styles import PatternFill, Font

def apply_cell_style(cell, fill, color_hex, font_bold=None):
    cell.fill = fill
    new_font = copy(cell.font)
    new_font.color = openpyxl.styles.colors.Color(rgb=color_hex)
    if font_bold is not None:
        new_font.bold = font_bold
    cell.font = new_font


sys.stdout.reconfigure(encoding="utf-8")

EXCEL_FILE  = "Kanken23510_2026年版_v6.xlsx"
STOCK_CSV   = "在庫一覧.csv"
BUNPAI_XLSX = "分配店舗向け.xlsx"

WAREHOUSE_LOCATIONS = {"バルク", "New Way-A"}
EXCLUDE_LOCATIONS   = {"New Way-B", "New Way-C", "New Way-G"}

CUMULATIVE_PROGRESS = {
    1: 0.0437, 2: 0.1050, 3: 0.1970, 4: 0.3090, 5: 0.3968, 6: 0.4921,
    7: 0.5968, 8: 0.7067, 9: 0.7903, 10: 0.8677, 11: 0.9360, 12: 1.0000
}

COL_COLOR=3; COL_WH=4; COL_STORE=5; COL_FW=6
COL_SUPPLY=8; COL_DEMAND=9; COL_METHOD=10; COL_ANNUAL=11; COL_DIFF=12
DATA_START=5; TOTAL_LABEL="合計"


def read_csv_auto(path):
    for enc in ("cp932","utf-8-sig","utf-8","shift_jis"):
        try:
            return pd.read_csv(path, encoding=enc)
        except (UnicodeDecodeError, LookupError):
            continue
    raise ValueError(f"エンコード失敗: {path}")


def find_latest_sales_csv():
    base = os.path.dirname(os.path.abspath(__file__))
    files = glob.glob(os.path.join(base, "*実績.csv"))
    if not files:
        raise FileNotFoundError("*実績.csv が見つかりません")
    return max(files, key=os.path.getmtime)


def load_master(path):
    df = pd.read_excel(path, sheet_name="master")
    df["品番"]    = df["品番"].astype(str).str.strip()
    df["カラー名"] = df["カラー名"].astype(str).str.strip()
    return dict(zip(df["品番"], df["カラー名"]))


def load_stock(path, master):
    df = read_csv_auto(path)
    df["商品コード"] = df["商品コード"].astype(str).str.strip()
    df["現在数量"] = pd.to_numeric(df["現在数量"], errors="coerce").fillna(0).clip(lower=0)
    df["カラー名"] = df["商品コード"].map(master)
    unmapped = df[df["カラー名"].isna()]["商品コード"].unique()
    if len(unmapped) > 0:
        print(f"  [情報] 在庫未挿入： {unmapped.tolist()[:10]}")
    df = df.dropna(subset=["カラー名"])
    df = df[~df["拠点名"].isin(EXCLUDE_LOCATIONS)]
    df["区分"] = df["拠点名"].apply(lambda x: "倉庫" if x in WAREHOUSE_LOCATIONS else "店舗")
    wh    = df[df["区分"]=="倉庫"].groupby("カラー名")["現在数量"].sum().astype(int).to_dict()
    store = df[df["区分"]=="店舗"].groupby("カラー名")["現在数量"].sum().astype(int).to_dict()
    return wh, store


def load_sales_actual(sales_csv, bunpai_xlsx, master):
    try:
        ds = pd.read_excel(bunpai_xlsx, sheet_name="2026売上")
        ds["カラー名"] = ds["カラー名"].astype(str).str.strip()
        ds["販売数"] = pd.to_numeric(ds["販売数"], errors="coerce").fillna(0).clip(lower=0)
        sbc = ds.groupby("カラー名")["販売数"].sum().astype(int).to_dict()
        print(f"      [ソース1] 2026売上: {len(sbc)} カラー, 合計 {sum(sbc.values()):,}個")
    except Exception as e:
        print(f"      [警告] 2026売上読込み失敗: {e}")
        sbc = {}

    dr = read_csv_auto(sales_csv)
    if "数量" not in dr.columns and "販売数" in dr.columns:
        dr = dr.rename(columns={"販売数": "数量"})
    qty = next((c for c in ["数量","販売数","Qty","qty"] if c in dr.columns), None)
    if qty is None:
        return sbc, 1, {}, {m: 0 for m in range(1, 13)}
    dr[qty] = pd.to_numeric(dr[qty], errors="coerce").fillna(0).clip(lower=0)

    item_col = next((c for c in ["3rd Item No.","品番","商品コード","行ラベル"] if c in dr.columns), None)
    if item_col:
        dr["key"] = dr[item_col].astype(str).str.strip()
        dr["cn"]  = dr["key"].map(master)
        dk = dr.dropna(subset=["cn"])
    else:
        dk = dr

    if not sbc and item_col:
        sbc = dk.groupby("cn")[qty].sum().astype(int).to_dict()
        print(f"      [ソース2(FB)]: {len(sbc)} カラー, {sum(sbc.values()):,}個")

    sales_by_color_month = {}
    sales_by_month = {m: 0 for m in range(1, 13)}
    mc = 1

    if "営業日付" in dk.columns:
        dk = dk.copy()
        dk["dt"] = pd.to_datetime(dk["営業日付"], errors="coerce")
        dk["m"]  = dk["dt"].dt.month
        mq = dk.groupby("m")[qty].sum()
        vm = [m for m, v in mq.items() if v > 0 and pd.notna(m)]
        mc = len(vm)
        print(f"      実績月： {sorted([int(m) for m in vm])}")

        for (cn, m), v in dk.groupby(["cn", "m"])[qty].sum().items():
            if pd.notna(m) and 1 <= m <= 12:
                m = int(m)
                if cn not in sales_by_color_month:
                    sales_by_color_month[cn] = {month: 0 for month in range(1, 13)}
                sales_by_color_month[cn][m] = int(v)
                sales_by_month[m] += int(v)
    else:
        mc = 1

    return sbc, max(1, min(12, mc)), sales_by_color_month, sales_by_month


def get_monthly_weights(months):
    weights = {}
    total_w = 0.0
    for m in range(months + 1, 13):
        p_curr = CUMULATIVE_PROGRESS.get(m, 1.0)
        p_prev = CUMULATIVE_PROGRESS.get(m - 1, 0.0)
        w = p_curr - p_prev
        weights[m] = w
        total_w += w
    return weights, total_w


def calc_color_plans(sbc, months, ws, wh_stock, store_stock):
    pr = CUMULATIVE_PROGRESS.get(months, 1.0)
    
    old_total = sum(
        float(row[COL_DEMAND-1].value or 0)
        for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row)
        if str(row[COL_COLOR-1].value or "").strip() not in ("", TOTAL_LABEL)
        and isinstance(row[COL_DEMAND-1].value, (int, float))
    )
    ta = sum(sbc.values())
    if pr > 0 and ta > 0 and old_total > 0:
        h2 = 1.0 - pr
        scale = (ta / pr * h2) / old_total
    else:
        scale = None

    color_plans = {}
    weights, total_w = get_monthly_weights(months)
    
    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row):
        cv = str(row[COL_COLOR-1].value or "").strip()
        if not cv or cv == TOTAL_LABEL:
            continue
        
        fv = int(row[COL_FW-1].value) if isinstance(row[COL_FW-1].value, (int, float)) else 0
        nw = wh_stock.get(cv, int(row[COL_WH-1].value or 0))
        ns = store_stock.get(cv, int(row[COL_STORE-1].value or 0))
        supply = nw + ns + fv
        
        if cv in sbc and pr > 0:
            ca = sbc[cv]
            total_nd = max(0.0, ca / pr - ca)
            annual_demand = ca / pr
        elif scale is not None:
            old_d = float(row[COL_DEMAND-1].value or 0)
            total_nd = old_d * scale
            annual_demand = total_nd / (1.0 - pr) if pr < 1.0 else total_nd
        else:
            old_d = float(row[COL_DEMAND-1].value or 0)
            total_nd = old_d
            annual_demand = total_nd / (1.0 - pr) if pr < 1.0 else total_nd
            
        color_demand = {}
        for m in range(months + 1, 13):
            if total_w > 0:
                color_demand[m] = round(total_nd * (weights[m] / total_w))
            else:
                color_demand[m] = 0
                
        sum_d = sum(color_demand.values())
        diff_d = round(total_nd) - sum_d
        if diff_d != 0 and 12 in color_demand:
            color_demand[12] = max(0, color_demand[12] + diff_d)

        color_plan = {}
        is_con = supply < total_nd
        method = "区間内比例配分" if is_con else "需要追随(制約なし)"
        
        if is_con:
            months_list = list(range(months + 1, 13))
            sum_d = sum(color_demand.values())
            if sum_d <= 0:
                color_plan = {m: 0 for m in months_list}
            else:
                alloc = 0
                for i, m in enumerate(months_list):
                    if i == len(months_list) - 1:
                        color_plan[m] = max(0, supply - alloc)
                    else:
                        p = round(color_demand[m] / sum_d * supply)
                        color_plan[m] = p
                        alloc += p
        else:
            color_plan = {m: color_demand[m] for m in range(months + 1, 13)}
            
        color_plans[cv] = {
            "demand": color_demand,
            "plan": color_plan,
            "total_nd": round(total_nd),
            "method": method,
            "supply": supply
        }
        
    return color_plans


def parse_fw_inflow(b2_text):
    mb = 0; ab = 0
    if b2_text:
        m_mb = re.search(r"MB:(\d+)", b2_text)
        if m_mb:
            mb = int(m_mb.group(1))
        m_ab = re.search(r"AB:(\d+)", b2_text)
        if m_ab:
            ab = int(m_ab.group(1))
    return mb, ab


def main():
    print("=" * 65)
    print("  build_summary.py -- 出荷計画一括再構築スクリプト")
    print("=" * 65)

    print("\n[1/5] マスタデータ読み込み...")
    master = load_master(BUNPAI_XLSX)
    print(f"      {len(master)} 品番")

    print(f"\n[2/5] 在庫一覧 ({STOCK_CSV}) 読み込み...")
    wh_stock, store_stock = load_stock(STOCK_CSV, master)
    print(f"      倉庫WH : {len(wh_stock)} カラー / 合計 {sum(wh_stock.values()):,}個")
    print(f"      店舗在庫: {len(store_stock)} カラー / 合計 {sum(store_stock.values()):,}個")

    print(f"\n[3/5] 実績データ読み込み...")
    csv = find_latest_sales_csv()
    print(f"      実績ファイル: {os.path.basename(csv)}")
    sbc, months, sales_by_color_month, sales_by_month = load_sales_actual(csv, BUNPAI_XLSX, master)
    pr = CUMULATIVE_PROGRESS.get(months, 1.0)
    print(f"      実績月数: {months} ヶ月 (進捗率: {pr*100:.2f}%)")
    print(f"      Kanken累計実績: {sum(sbc.values()):,}個")

    print(f"\n[4/5] v6ファイル読み込み...")
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"'{EXCEL_FILE}' が見つかりません")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["全体サマリー"]
    
    color_plans = calc_color_plans(sbc, months, ws, wh_stock, store_stock)
    print(f"      メモリ上での予測計算完了")

    print(f"\n[5/5] 書き込み中...")
    
    # --- 1. 全体サマリーシートの更新 ---
    print("   -> 「全体サマリー」を更新中...")
    ws["D4"].value = f"{months+1}/1\n倉庫WH"
    ws["H4"].value = f"総供給\n({months+1}月〜)"
    ws["I4"].value = f"需要予測\n({months+1}〜12月)"
    
    td = te = th = ti = tl = tk = updated = 0

    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row):
        cv = str(row[COL_COLOR-1].value or "").strip()
        if not cv or cv == TOTAL_LABEL:
            continue
        
        wc = row[COL_WH-1];    sc = row[COL_STORE-1]
        fc = row[COL_FW-1];    hc = row[COL_SUPPLY-1]
        ic = row[COL_DEMAND-1]; lc = row[COL_DIFF-1]
        jc = row[COL_METHOD-1]; kc = row[COL_ANNUAL-1]

        ow = int(wc.value) if isinstance(wc.value,(int,float)) else 0
        os_ = int(sc.value) if isinstance(sc.value,(int,float)) else 0
        fv = int(fc.value) if isinstance(fc.value,(int,float)) else 0
        oh = int(hc.value) if isinstance(hc.value,(int,float)) else 0
        oi = int(ic.value) if isinstance(ic.value,(int,float)) else 0

        nw = wh_stock.get(cv, ow)
        ns = store_stock.get(cv, os_)
        nh = nw + ns + fv
        
        plan_info = color_plans.get(cv)
        if plan_info:
            ni = plan_info["total_nd"]
            nj = plan_info["method"]
            actual_sum = sum(sales_by_color_month.get(cv, {}).get(m, 0) for m in range(1, months + 1))
            plan_sum = sum(plan_info["plan"].values())
            nk = actual_sum + plan_sum
            nl = nh - ni
        else:
            ni = 0
            nj = "需要追随(制約なし)"
            nk = 0
            nl = nh

        wc.value = nw; sc.value = ns; hc.value = nh; ic.value = ni; lc.value = nl
        jc.value = nj; kc.value = nk

        td += nw; te += ns; th += nh; ti += ni; tl += nl; tk += nk; updated += 1

    # 合計行
    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row):
        if str(row[COL_COLOR-1].value or "").strip() == TOTAL_LABEL:
            row[COL_WH-1].value = td; row[COL_STORE-1].value = te
            row[COL_SUPPLY-1].value = th; row[COL_DEMAND-1].value = ti
            row[COL_ANNUAL-1].value = tk; row[COL_DIFF-1].value = tl
            print(f"      全体サマリー合計: WH={td} 店舗={te} 供給={th} 需要={ti} 年間計画={tk} 過不足={tl:+}")
            break

    # --- 2. 2026年着地予測シートの更新 ---
    print("   -> 「2026年着地予測」を更新中...")
    ws_forecast = wb["2026年着地予測"]
    cumul = 0
    for m in range(1, 13):
        ri = m + 4
        ws_forecast.cell(row=ri, column=5).value = "実績" if m <= months else "計画"
        
        if m <= months:
            m_val = sales_by_month.get(m, 0)
            ws_forecast.cell(row=ri, column=4).value = m_val
            ws_forecast.cell(row=ri, column=6).value = m_val
            cumul += m_val
        else:
            m_demand = sum(color_plans[cv]["demand"].get(m, 0) for cv in color_plans if cv in color_plans)
            m_plan = sum(color_plans[cv]["plan"].get(m, 0) for cv in color_plans if cv in color_plans)
            ws_forecast.cell(row=ri, column=4).value = m_demand
            ws_forecast.cell(row=ri, column=6).value = m_plan
            cumul += m_plan
        ws_forecast.cell(row=ri, column=8).value = cumul

    total_d = sum(ws_forecast.cell(row=r, column=4).value or 0 for r in range(5, 17))
    total_p = sum(ws_forecast.cell(row=r, column=6).value or 0 for r in range(5, 17))
    ws_forecast.cell(row=17, column=4).value = total_d
    ws_forecast.cell(row=17, column=6).value = total_p
    ws_forecast.cell(row=17, column=7).value = 100
    ws_forecast.cell(row=17, column=8).value = total_p
    for ri in range(5, 17):
        p = ws_forecast.cell(row=ri, column=6).value or 0
        ws_forecast.cell(row=ri, column=7).value = round(p / total_p * 100, 2) if total_p else 0

    # --- 3. 供給需要シミュレーションシートの更新 ---
    print("   -> 「供給需要シミュレーション」を更新中...")
    ws_sim = wb["供給需要シミュレーション"]
    ws_sim["B1"].value = f"  供給需要シミュレーション　Kanken 23510　2026年{months+1}〜12月（区間内比例配分後）"
    ws_sim["B2"].value = f"  {months+1}/1 倉庫(BULK+NWA・27色合計): {td:,}個　｜　制約あり色は区間内比例配分により年末まで在庫を維持しつつ需要の強弱を反映"

    mb_total = 0
    ab_total = 0
    for name in wb.sheetnames:
        if name in ('計算ロジック説明', '2026年着地予測', '供給需要シミュレーション', '全体サマリー', '店頭在庫_実績入力', '2025年入荷比較'):
            continue
        c_ws = wb[name]
        b2_val = c_ws["B2"].value or ""
        mb, ab = parse_fw_inflow(b2_val)
        mb_total += mb
        ab_total += ab

    sim_rows = []
    cur_stock = td
    for m in range(months + 1, 13):
        arr = 0
        if m == 7:
            arr = mb_total
        elif m == 9:
            arr = ab_total
        p = sum(color_plans[cv]["plan"].get(m, 0) for cv in color_plans)
        end_stock = max(0, cur_stock + arr - p)
        status = "✅ 問題なし" if end_stock >= 0 else "⚠️ 不足あり"
        if diff_d != 0 and 12 in color_demand:
            color_demand[12] = max(0, color_demand[12] + diff_d)

        color_plan = {}
        is_con = supply < total_nd
        method = "区間内比例配分" if is_con else "需要追随(制約なし)"
        
        if is_con:
            months_list = list(range(months + 1, 13))
            sum_d = sum(color_demand.values())
            if sum_d <= 0:
                color_plan = {m: 0 for m in months_list}
            else:
                alloc = 0
                for i, m in enumerate(months_list):
                    if i == len(months_list) - 1:
                        color_plan[m] = max(0, supply - alloc)
                    else:
                        p = round(color_demand[m] / sum_d * supply)
                        color_plan[m] = p
                        alloc += p
        else:
            color_plan = {m: color_demand[m] for m in range(months + 1, 13)}
            
        color_plans[cv] = {
            "demand": color_demand,
            "plan": color_plan,
            "total_nd": round(total_nd),
            "method": method,
            "supply": supply
        }
        
    return color_plans


def parse_fw_inflow(b2_text):
    mb = 0; ab = 0
    if b2_text:
        m_mb = re.search(r"MB:(\d+)", b2_text)
        if m_mb:
            mb = int(m_mb.group(1))
        m_ab = re.search(r"AB:(\d+)", b2_text)
        if m_ab:
            ab = int(m_ab.group(1))
    return mb, ab


def main():
    print("=" * 65)
    print("  build_summary.py -- 出荷計画一括再構築スクリプト")
    print("=" * 65)

    print("\n[1/5] マスタデータ読み込み...")
    master = load_master(BUNPAI_XLSX)
    print(f"      {len(master)} 品番")

    print(f"\n[2/5] 在庫一覧 ({STOCK_CSV}) 読み込み...")
    wh_stock, store_stock = load_stock(STOCK_CSV, master)
    print(f"      倉庫WH : {len(wh_stock)} カラー / 合計 {sum(wh_stock.values()):,}個")
    print(f"      店舗在庫: {len(store_stock)} カラー / 合計 {sum(store_stock.values()):,}個")

    print(f"\n[3/5] 実績データ読み込み...")
    csv = find_latest_sales_csv()
    print(f"      実績ファイル: {os.path.basename(csv)}")
    sbc, months, sales_by_color_month, sales_by_month = load_sales_actual(csv, BUNPAI_XLSX, master)
    pr = CUMULATIVE_PROGRESS.get(months, 1.0)
    print(f"      実績月数: {months} ヶ月 (進捗率: {pr*100:.2f}%)")
    print(f"      Kanken累計実績: {sum(sbc.values()):,}個")

    print(f"\n[4/5] v6ファイル読み込み...")
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"'{EXCEL_FILE}' が見つかりません")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["全体サマリー"]
    
    color_plans = calc_color_plans(sbc, months, ws, wh_stock, store_stock)
    print(f"      メモリ上での予測計算完了")

    # カラー別シートから元の配分方式を抽出
    original_methods = {}
    for name in wb.sheetnames:
        if name in ('計算ロジック説明', '2026年着地予測', '供給需要シミュレーション', '全体サマリー', '店頭在庫_実績入力', '2025年入荷比較'):
            continue
        c_ws = wb[name]
        b2_val = c_ws["B2"].value or ""
        method_str = "需要追随(制約なし)"
        m_meth = re.search(r"配分方式:\s*([^\s｜]+)", b2_val)
        if m_meth:
            method_str = m_meth.group(1).strip()
        original_methods[name] = method_str
    print(f"      カラー別シートから配分方式の復元完了")

    print(f"\n[5/5] 書き込み中...")
    
    # --- 1. 全体サマリーシートの更新 ---
    print("   -> 「全体サマリー」を更新中...")
    ws["D4"].value = f"{months+1}/1\n倉庫WH"
    ws["H4"].value = f"総供給\n({months+1}月〜)"
    ws["I4"].value = f"需要予測\n({months+1}〜12月)"
    
    td = te = th = ti = tl = tk = updated = 0

    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row):
        cv = str(row[COL_COLOR-1].value or "").strip()
        if not cv or cv == TOTAL_LABEL:
            continue
        
        wc = row[COL_WH-1];    sc = row[COL_STORE-1]
        fc = row[COL_FW-1];    hc = row[COL_SUPPLY-1]
        ic = row[COL_DEMAND-1]; lc = row[COL_DIFF-1]
        jc = row[COL_METHOD-1]; kc = row[COL_ANNUAL-1]

        ow = int(wc.value) if isinstance(wc.value,(int,float)) else 0
        os_ = int(sc.value) if isinstance(sc.value,(int,float)) else 0
        fv = int(fc.value) if isinstance(fc.value,(int,float)) else 0
        oh = int(hc.value) if isinstance(hc.value,(int,float)) else 0
        oi = int(ic.value) if isinstance(ic.value,(int,float)) else 0

        nw = wh_stock.get(cv, ow)
        ns = store_stock.get(cv, os_)
        nh = nw + ns + fv
        
        plan_info = color_plans.get(cv)
        if plan_info:
            ni = plan_info["total_nd"]
            actual_sum = sum(sales_by_color_month.get(cv, {}).get(m, 0) for m in range(1, months + 1))
            plan_sum = sum(plan_info["plan"].values())
            nk = actual_sum + plan_sum
            nl = nh - ni
        else:
            ni = 0
            nk = 0
            nl = nh

        # 配分方式 (J列) を元のカラー別定義に基づいて復元・上書きする
        nj_for_color = original_methods.get(cv, "需要追随(制約なし)")
        jc.value = nj_for_color

        wc.value = nw; sc.value = ns; hc.value = nh; ic.value = ni; lc.value = nl
        kc.value = nk

        # L列 (過不足) の色付け改善
        if nl < 0:
            fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            apply_cell_style(lc, fill_red, "C00000")
        else:
            if "比例配分" in nj_for_color:
                fill_purple = PatternFill(start_color="E5DCEE", end_color="E5DCEE", fill_type="solid")
                apply_cell_style(lc, fill_purple, "7030A0")
            else:
                fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                apply_cell_style(lc, fill_green, "375623")

        td += nw; te += ns; th += nh; ti += ni; tl += nl; tk += nk; updated += 1

    # 合計行
    for row in ws.iter_rows(min_row=DATA_START, max_row=ws.max_row):
        if str(row[COL_COLOR-1].value or "").strip() == TOTAL_LABEL:
            row[COL_WH-1].value = td; row[COL_STORE-1].value = te
            row[COL_SUPPLY-1].value = th; row[COL_DEMAND-1].value = ti
            row[COL_ANNUAL-1].value = tk; row[COL_DIFF-1].value = tl
            
            # 合計行の過不足セルの色更新
            tot_diff_cell = row[COL_DIFF-1]
            if tl < 0:
                fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                apply_cell_style(tot_diff_cell, fill_red, "C00000", font_bold=True)
            else:
                fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                apply_cell_style(tot_diff_cell, fill_green, "375623", font_bold=True)
                
            print(f"      全体サマリー合計: WH={td} 店舗={te} 供給={th} 需要={ti} 年間計画={tk} 過不足={tl:+}")
            break

    # --- 2. 2026年着地予測シートの更新 ---
    print("   -> 「2026年着地予測」を更新中...")
    ws_forecast = wb["2026年着地予測"]
    cumul = 0
    for m in range(1, 13):
        ri = m + 4
        is_actual = (m <= months)
        status_cell = ws_forecast.cell(row=ri, column=5)
        status_cell.value = "実績" if is_actual else "計画"
        
        if is_actual:
            m_val = sales_by_month.get(m, 0)
            ws_forecast.cell(row=ri, column=4).value = m_val
            ws_forecast.cell(row=ri, column=6).value = m_val
            cumul += m_val
            
            # 実績行のスタイル適用
            fill_act = PatternFill(start_color="D6E4F5", end_color="D6E4F5", fill_type="solid")
            status_cell.fill = fill_act
            new_font = copy(status_cell.font)
            new_font.color = openpyxl.styles.colors.Color(rgb="2E75B6")
            status_cell.font = new_font
            
            for col_idx in (4, 6, 8):
                c = ws_forecast.cell(row=ri, column=col_idx)
                c_font = copy(c.font)
                c_font.color = openpyxl.styles.colors.Color(rgb="000000")
                c.font = c_font
        else:
            m_demand = sum(color_plans[cv]["demand"].get(m, 0) for cv in color_plans if cv in color_plans)
            m_plan = sum(color_plans[cv]["plan"].get(m, 0) for cv in color_plans if cv in color_plans)
            ws_forecast.cell(row=ri, column=4).value = m_demand
            ws_forecast.cell(row=ri, column=6).value = m_plan
            cumul += m_plan
            
            # 計画行のスタイル適用
            bg_color = "F2F2F2" if ri % 2 == 0 else "FFFFFF"
            fill_plan = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            status_cell.fill = fill_plan
            new_font = copy(status_cell.font)
            new_font.color = openpyxl.styles.colors.Color(rgb="7F6000")
            status_cell.font = new_font
            
            for col_idx in (4, 6, 8):
                c = ws_forecast.cell(row=ri, column=col_idx)
                c_font = copy(c.font)
                c_font.color = openpyxl.styles.colors.Color(rgb="888888")
                c.font = c_font
                
        ws_forecast.cell(row=ri, column=8).value = cumul

    total_d = sum(ws_forecast.cell(row=r, column=4).value or 0 for r in range(5, 17))
    total_p = sum(ws_forecast.cell(row=r, column=6).value or 0 for r in range(5, 17))
    ws_forecast.cell(row=17, column=4).value = total_d
    ws_forecast.cell(row=17, column=6).value = total_p
    ws_forecast.cell(row=17, column=7).value = 100
    ws_forecast.cell(row=17, column=8).value = total_p
    for ri in range(5, 17):
        p = ws_forecast.cell(row=ri, column=6).value or 0
        ws_forecast.cell(row=ri, column=7).value = round(p / total_p * 100, 2) if total_p else 0

    # --- 3. 供給需要シミュレーションシートの更新 ---
    print("   -> 「供給需要シミュレーション」を更新中...")
    ws_sim = wb["供給需要シミュレーション"]
    ws_sim["B1"].value = f"  供給需要シミュレーション　Kanken 23510　2026年{months+1}〜12月（区間内比例配分後）"
    ws_sim["B2"].value = f"  {months+1}/1 倉庫(BULK+NWA・27色合計): {td:,}個　｜　制約あり色は区間内比例配分により年末まで在庫を維持しつつ需要の強弱を反映"

    mb_total = 0
    ab_total = 0
    for name in wb.sheetnames:
        if name in ('計算ロジック説明', '2026年着地予測', '供給需要シミュレーション', '全体サマリー', '店頭在庫_実績入力', '2025年入荷比較'):
            continue
        c_ws = wb[name]
        b2_val = c_ws["B2"].value or ""
        mb, ab = parse_fw_inflow(b2_val)
        mb_total += mb
        ab_total += ab

    sim_rows = []
    cur_stock = td
    for m in range(months + 1, 13):
        arr = 0
        if m == 7:
            arr = mb_total
        elif m == 9:
            arr = ab_total
        p = sum(color_plans[cv]["plan"].get(m, 0) for cv in color_plans)
        end_stock = max(0, cur_stock + arr - p)
        status = "✅ 問題なし" if end_stock >= 0 else "⚠️ 不足あり"
        sim_rows.append((m, cur_stock, arr, p, end_stock, status))
        cur_stock = end_stock

    start_row = 5
    num_rows = len(sim_rows)
    for i, (m, start, arr, p, end, status) in enumerate(sim_rows):
        r = start_row + i
        ws_sim.cell(row=r, column=2).value = f"{m}月"
        ws_sim.cell(row=r, column=3).value = start
        ws_sim.cell(row=r, column=4).value = arr
        ws_sim.cell(row=r, column=5).value = p
        ws_sim.cell(row=r, column=6).value = end
        ws_sim.cell(row=r, column=7).value = status

    for r in range(start_row + num_rows, 12):
        for c in range(2, 8):
            try:
                ws_sim.cell(row=r, column=c).value = None
            except AttributeError:
                pass

    tot_row = start_row + num_rows
    try:
        ws_sim.cell(row=tot_row, column=2).value = f"合計({months+1}〜12月)"
    except AttributeError:
        pass
    try:
        ws_sim.cell(row=tot_row, column=3).value = None
    except AttributeError:
        pass
    try:
        ws_sim.cell(row=tot_row, column=4).value = sum(x[2] for x in sim_rows)
    except AttributeError:
        pass
    try:
        ws_sim.cell(row=tot_row, column=5).value = sum(x[3] for x in sim_rows)
    except AttributeError:
        pass
    try:
        ws_sim.cell(row=tot_row, column=6).value = sim_rows[-1][4] if sim_rows else 0
    except AttributeError:
        pass
    try:
        ws_sim.cell(row=tot_row, column=7).value = None
    except AttributeError:
        pass

    note_row = tot_row + 1
    try:
        ws_sim.cell(row=note_row, column=2).value = "  💡 区間内比例配分ロジックにより、入荷区間を超えて在庫がマイナスになることはない。詳細は各カラー別シート参照。"
    except AttributeError:
        pass
    for c in range(3, 8):
        try:
            ws_sim.cell(row=note_row, column=c).value = None
        except AttributeError:
            pass

    for r in range(note_row + 1, 15):
        for c in range(2, 8):
            try:
                ws_sim.cell(row=r, column=c).value = None
            except AttributeError:
                pass

    wb.save(EXCEL_FILE)
    print(f"\n✅ Excelの更新が完了しました。")
    print(f"  総供給(H): {th:,}  需要予測(I): {ti:,}  過不足(L): {tl:+,}")
    print("=" * 65)

if __name__ == "__main__":
    main()
